
import logging
import openpyxl
import os
import math

logger = logging.getLogger(__name__)

EXCEL_PATH = r"c:/Users/Dell/Desktop/Rashad/Nurlan/Oil&Gas/blend/TEST TABLE.xlsx"

def get_available_tanks():
    """
    Reads the Excel file and returns a list of available tanks (Column J onwards)
    where 'IN TANK MTA' (Row 11) > 0.
    """
    if not os.path.exists(EXCEL_PATH):
        logger.error(f"Excel file not found at {EXCEL_PATH}")
        return []

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb['BLEND CALCULATION']
    except Exception as e:
        logger.error(f"Error reading Excel: {e}")
        return []

    tanks = []
    
    # Iterate columns from J (9) to AK (36)
    # openpyxl uses 1-based indexing. J is 10.
    # Actually let's find the range dynamically or hardcode as per analysis.
    # J=10, AK=37? Let's use numeric indices.
    
    # Row 11 is IN TANK MTA. 
    # Row 27 is Density.
    # Row 28 is Viscosity.
    # Row 31 is Sulphur.
    # Row 30 is Water.
    # Row 32 is Flash Point.
    # Row 33 is Pour Point.
    
    start_col = 10 # J
    end_col = 37   # AK
    
    for col_idx in range(start_col, end_col + 1):
        try:
            qty_cell = ws.cell(row=11, column=col_idx)
            qty = qty_cell.value
            
            if isinstance(qty, (int, float)) and qty > 0:
                name_cell = ws.cell(row=2, column=col_idx) # Tank Names usually in Row 2 or nearby?
                # In analysis: "Row 26... TANK 1 TANK 2". "Row 10... CAP MTA".
                # Let's check where the label is. 
                # Analysis says "Row 27... Col J: TANK 1". 
                # Let's use the explicit Tank Names from Row 26 or header.
                # Actually, in the `debug_formulas.py` output: "Col J: TANK 1" for Row 27.
                # Wait, Row 27 is Density. It has "TANK 1" as value?
                # Ah, "Col J: 0.94".
                # Where is "TANK 1" label?
                # In `analyze_excel.py` output: "Row 26... TANK 1 TANK 2".
                # So Header is likely Row 26.
                
                tank_name = ws.cell(row=26, column=col_idx).value or f"Tank {col_idx}"
                
                # Properties
                density = ws.cell(row=27, column=col_idx).value
                viscosity = ws.cell(row=28, column=col_idx).value
                water = ws.cell(row=29, column=col_idx).value # Wait, check row mapping
                
                # Re-verify row mapping from analysis
                # Row 27: Density
                # Row 28: Viscosity
                # Row 29: Water? No, analysis says Row 29 is Viscosity?
                # Let's check `debug_rows.py` output.
                # B28: Density at 15 oC.
                # B29: Kin. Viscosity at 50 deg C.
                # Ah, so Row 28 is Density, Row 29 is Viscosity.
                # Row 31: Sulphur.
                # Row 30: Water?
                # "Water by distillation ... Row 30" (from `analyze_excel.py` output, line ~400).
                # Yes.
                # Row 32: Flash Point.
                # Row 33: Pour Point.
                
                # Extract simple properties
                props = {
                    'density': density,
                    'viscosity': viscosity, # cSt
                    'water': ws.cell(row=30, column=col_idx).value,
                    'sulphur': ws.cell(row=31, column=col_idx).value,
                    'flash_point': ws.cell(row=32, column=col_idx).value,
                    'pour_point': ws.cell(row=33, column=col_idx).value
                }
                
                # Clean up None values
                props = {k: (float(v) if isinstance(v, (int, float)) else 0) for k, v in props.items()}
                
                tanks.append({
                    'id': f"col_{col_idx}",
                    'name': str(tank_name),
                    'quantity': float(qty),
                    'properties': props
                })
        except Exception as e:
            logger.warning(f"Error processing column {col_idx}: {e}")
            continue

    return tanks

def get_target_specs():
    """
    Reads Column E for target specs.
    """
    if not os.path.exists(EXCEL_PATH):
        return {}
    
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['BLEND CALCULATION']
    
    # Specs in Column E (5)
    specs = {
        'density_max': ws.cell(row=28, column=5).value, # Row 28 is Density
        'viscosity_max': ws.cell(row=29, column=5).value, # Row 29 is Visc
        'water_max': ws.cell(row=30, column=5).value,
        'sulphur_max': ws.cell(row=31, column=5).value,
        'flash_point_min': ws.cell(row=32, column=5).value,
        'pour_point_max': ws.cell(row=33, column=5).value,
    }
    
    # Clean up
    clean_specs = {}
    for k, v in specs.items():
        if isinstance(v, (int, float)):
            clean_specs[k] = float(v)
            
    return clean_specs
