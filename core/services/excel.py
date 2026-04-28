import logging
import openpyxl
import os
from django.conf import settings

logger = logging.getLogger(__name__)

EXCEL_PATH = os.path.join(settings.BASE_DIR, 'uploads', 'TEST TABLE.xlsx')


def get_blending_data(included_terminals=None, excluded_terminals=None):
    """
    Reads one or more Excel files to extract:
    1. Tanks: Scanned from Row 27 (Cols J onwards).
    2. Properties/Specs: Scanned from Rows 19-81 (Cols B & E).
    3. Property Matrix: Values for each tank for each property.
    """
    from ..models import Terminal
    
    # 1. Gather files to process
    terminals = Terminal.objects.all()
    
    if included_terminals:
        terminals = terminals.filter(name__in=included_terminals)
    if excluded_terminals:
        terminals = terminals.exclude(name__in=excluded_terminals)
        
    process_list = []
    if terminals.exists():
        for t in terminals:
            if t.excel_file and os.path.exists(t.excel_file.path):
                process_list.append({'path': t.excel_file.path, 'name': t.name})
    else:
        # Fallback to legacy path if no terminals defined
        if os.path.exists(EXCEL_PATH):
            process_list.append({'path': EXCEL_PATH, 'name': 'Default'})

    if not process_list:
        logger.error("No valid blending source files found.")
        return None

    data = {
        'tanks': [],
        'properties': []
    }
    
    # Track property metadata to merge efficiently
    prop_map = {} # clean_name -> {spec, is_spec, safety_margin, tank_values}

    # Load Safety Margins from JSON
    margins = {}
    margin_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'safety_margins.json')
    if os.path.exists(margin_file):
        try:
            import json
            with open(margin_file, 'r') as f:
                margins = json.load(f)
        except Exception as e:
            logger.error(f"Error loading safety margins: {e}")

    for item in process_list:
        path = item['path']
        t_name = item['name']
        
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb['BLEND CALCULATION']
        except Exception as e:
            logger.error(f"Error reading Excel {path}: {e}")
            continue

        # A. Identify Tanks in this file
        tank_cols = []
        for col_idx in range(10, 50):
            tank_name = ws.cell(row=27, column=col_idx).value
            if tank_name:
                qty_val = ws.cell(row=13, column=col_idx).value
                price_val = ws.cell(row=80, column=col_idx).value
                
                try:
                    qty = float(qty_val) if qty_val is not None else 0.0
                except: qty = 0.0
                
                try:
                    price = float(price_val) if price_val is not None else 0.0
                except: price = 0.0

                tank_id = f"{t_name}_col_{col_idx}"
                tank_data = {
                    'id': tank_id,
                    'name': f"{tank_name} ({t_name})",
                    'terminal': t_name,
                    'quantity': qty,
                    'price': price,
                    'col_idx': col_idx,
                    'path': path
                }
                data['tanks'].append(tank_data)
                tank_cols.append({'id': tank_id, 'col_idx': col_idx})

        # B. Extract Properties in this file
        for row_idx in range(19, 82):
            if row_idx == 27: continue 
            prop_name_val = ws.cell(row=row_idx, column=2).value # Col B
            spec_val = ws.cell(row=row_idx, column=5).value  # Col E
            
            if not prop_name_val: continue
            
            clean_name = str(prop_name_val).strip()
            is_spec = isinstance(spec_val, (int, float))
            
            if clean_name not in prop_map:
                prop_map[clean_name] = {
                    'name': clean_name,
                    'spec': float(spec_val) if is_spec else None,
                    'is_spec': is_spec,
                    'safety_margin': margins.get(clean_name, 0.0),
                    'tank_values': {}
                }
            
            # Map tank values for tanks FOUND IN THIS FILE
            for tc in tank_cols:
                val = ws.cell(row=row_idx, column=tc['col_idx']).value
                try:
                    f_val = float(val) if val is not None else 0.0
                except:
                    f_val = 0.0
                prop_map[clean_name]['tank_values'][tc['id']] = f_val

    # Convert Map to List
    data['properties'] = list(prop_map.values())
    
    # Ensure all tanks have a value for every property (fill with 0 if missing from a specific file)
    for p in data['properties']:
        for t in data['tanks']:
            if t['id'] not in p['tank_values']:
                p['tank_values'][t['id']] = 0.0

    return data
